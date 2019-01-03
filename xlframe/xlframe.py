import datetime as _dt
import os as _os
from copy import copy as _copy
from itertools import count as _count

import pandas as _pd
from openpyxl import load_workbook as _load_workbook
from openpyxl.styles.builtins import styles as _styles
from openpyxl.utils import get_column_letter as _get_column_letter
from openpyxl.worksheet import table as _table
from openpyxl.worksheet.hyperlink import Hyperlink as _Hyperlink

from . import utils as _utils
from .style import Style as _Style

__all__ = ['XlFrame']
_no_time = _dt.time(0)
_filler = {'fill_pattern': _utils.FillPattern.solid}


class XlFrame:
    utils = _utils

    def __init__(self, dataframe, style=None, header_style=None, index_style=None, *, number_style=None,
                 date_style=None, datetime_style=None, timedelta_style=None, use_default_formats=True):
        """
        Class for styling dataframes. Styling based around openpyxl's NamedStyle.
        xlframe.Style available as an alternative to using NamedStyle.

        Pandas-like syntax for assigning styles.

        Assign to column/s:
            my_xlframe['Column1'] = my_style
            my_xlframe[['Column1', 'Column2']] = my_style

        Use .styles or .istyles property like you would pandas.DataFrame.loc or .iloc but to assign styles:
            my_xlframe.styles[2:10, ['Column3', 'Column4']] = my_style
            my_xlframe.istyles[2:10, [2, 3]] = my_style

        To edit the existing style at a location pass a dictionary of changes:
            my_xlframe.styles[2:10, ['Column3', 'Column4']] = {'bold': True, 'font_color': 'red'}
        Accepts any kwargs xlframe.Style does.

        .auto_fit() attempts to fit column widths based on their contents.
        .to_excel() to export styled dataframe. Supports .xlsx and .xlsm file formats.

        Default type specific styling adjusts alignment and number format of style arg.

        :param dataframe: DateFrame to style.
        :type dataframe: pandas.DataFrame.
        :param style: Initial style to apply to frame. Default xlframe.Style.default_style().
        :type style: openpyxl.NamedStyle or xlframe.Style.
        :param header_style: Style for headers to use instead of style.
            Default center aligns, bolds and adds thin borders to style.
        :type header_style: openpyxl.NamedStyle or xlframe.Style.
        :param index_style: Style for index to use instead of style.
            Default bolds and adds thin borders to style.
            Alignment and number format based on index type.
        :type index_style: openpyxl.NamedStyle or xlframe.Style.
        :param number_style: Style for numeric columns to use instead of style.
            Default right aligns style and sets number format to utils.NumberFormats.general.
        :type number_style: openpyxl.NamedStyle or xlframe.Style.
        :param date_style: Style for date columns to use instead of style.
            Default right aligns style and sets number format to utils.Options.default_date_format.
        :type date_style: openpyxl.NamedStyle or xlframe.Style.
        :param datetime_style: Style for datetime columns to use instead of style.
            Default right aligns style and sets number format to utils.Options.default_datetime_format.
        :type datetime_style: openpyxl.NamedStyle or xlframe.Style.
        :param timedelta_style: Style for timedelta columns to use instead of style.
            Default sets number format to utils.Options.default_timedelta_format.
            Tries to align left or right depending on if number format is textual or not.
        :type timedelta_style: openpyxl.NamedStyle or xlframe.Style.
        :param use_default_formats: Apply default formatting where a style is not supplied.
        :type use_default_formats: Boolean.
        """
        if not isinstance(dataframe, _pd.DataFrame):
            raise TypeError('Must pass pandas.DataFrame.')
        if isinstance(dataframe.index, _pd.MultiIndex) or isinstance(dataframe.columns, _pd.MultiIndex):
            raise NotImplementedError('No support for pandas.MultiIndex on index or columns.')

        self.named_styles = dict()
        self._named_styles = _styles.copy()
        self.builtins = tuple(sorted(_styles))

        if not style:
            style = _Style.default_style()
        style = self._style_parser(style)
        index = index_style is None

        if use_default_formats:

            base_style = self._named_styles[style]

            index_style = self._style_parser(
                index_style if index_style else _Style._default_index_style(base_style)
            )
            header_style = self._style_parser(
                header_style if header_style else _Style._default_header_style(base_style)
            )

            if date_style is None:
                date_style = _Style._default_date_style(base_style)

            if number_style is None:
                number_style = _Style._default_number_style(base_style)

            if datetime_style is None:
                datetime_style = _Style._default_datetime_style(base_style)

            if timedelta_style is None:
                timedelta_style = _Style._default_timedelta_style(base_style)

            date_style = self._style_parser(date_style)
            number_style = self._style_parser(number_style)
            datetime_style = self._style_parser(datetime_style)
            timedelta_style = self._style_parser(timedelta_style)
        else:
            index_style = self._style_parser(index_style if index_style else style)
            header_style = self._style_parser(header_style if header_style else style)

        self.dataframe = self.df = dataframe
        self._styleframe = self._sf = _pd.DataFrame(
            data=style, index=self.dataframe.index, columns=self.dataframe.columns
        )

        self._index_styles = _pd.Series(data=index_style, index=dataframe.index, name='IndexStyles')
        self._header_styles = _pd.Series(data=header_style, index=dataframe.columns, name='HeaderStyles')

        self._row_heights = _pd.Series(
            data=float(_utils.Options.default_row_height), index=self.dataframe.index, name='RowHeights'
        )
        self._column_widths = _pd.Series(
            data=float(_utils.Options.default_column_width), index=self.dataframe.columns, name='ColumnWidths'
        )

        self._index_width = float(_utils.Options.default_column_width)
        self._header_height = float(_utils.Options.default_row_height)

        self._styleframe.style_loc = _StyleIndexer(self, self._styleframe.loc)
        self._styleframe.style_iloc = _StyleIndexer(self, self._styleframe.iloc)

        self._index_styles.style_idxr = _SeriesIndexer(self, self._index_styles)
        self._header_styles.style_idxr = _SeriesIndexer(self, self._header_styles)

        self._table_args = None
        self._hyperlinks = None
        self._slicer = _Slicer(self, 'loc')
        self._islicer = _Slicer(self, 'iloc')
        self._defaults_used = use_default_formats

        if number_style or date_style or datetime_style or timedelta_style:
            self._style_by_type(
                idxr=(slice(None), slice(None)),
                index=index and use_default_formats,
                default_style=style,
                number_style=number_style if number_style else style,
                date_style=date_style if date_style else style,
                datetime_style=datetime_style if datetime_style else style,
                timedelta_style=timedelta_style if timedelta_style else style,
            )

    def to_excel(self, excel_writer='output.xlsx', sheet_name='Sheet1', *, protect_sheet=False,
                 right_to_left=False, columns_to_hide=None, add_filters=False, replace_sheet=False,
                 auto_fit=None, **kwargs):
        """
        Export to excel. Supports .xlsx/.xlsm. See pandas.DataFrame.to_excel() for more.

        :param excel_writer: ExcelWriter or file path to export to.
        :type excel_writer: ExcelWriter or string.
        :param sheet_name: Sheet name to export to.
        :type sheet_name: string.
        :param protect_sheet: Enable protection on sheet.
        :type protect_sheet: boolean.
        :param right_to_left: Make sheet right to left oriented.
        :type right_to_left: boolean.
        :param columns_to_hide: Columns to make hidden.
        :type columns_to_hide: string, int or list-like.
        :param add_filters: Add excel filters to header row.
        :type add_filters: boolean.
        :param replace_sheet: If sheet_name already exists delete it first.
        :type replace_sheet: boolean.
        :param auto_fit: Columns to autofit. Can pass True to fit all columns.
        :type auto_fit: list-like or boolean.
        :param kwargs: Passed to pandas.DataFrame.to_excel().
        :return: pandas.ExcelWriter.
        """

        save = kwargs.pop('save', isinstance(excel_writer, str))
        # pandas.to_excel defaults
        index = kwargs.pop('index', True)
        header = kwargs.pop('header', True)
        columns = kwargs.pop('columns', None)
        engine = kwargs.pop('engine', 'openpyxl')
        startcol = kwargs.pop('startcol', 0)
        startrow = headerrow = kwargs.pop('startrow', 0)

        index_label = kwargs.get('index_label', self.index.name)

        if self._table_args and not header:
            raise ValueError('Cannot format as table without headers.')

        if columns is not None:
            return self.loc[:, columns].to_excel(
                excel_writer=excel_writer, sheet_name=sheet_name, protect_sheet=protect_sheet,
                right_to_left=right_to_left, columns_to_hide=columns_to_hide, add_filters=add_filters,
                replace_sheet=replace_sheet, auto_fit=auto_fit, header=header, index=index,
                startcol=startcol, startrow=startrow, engine=engine, save=save, **kwargs
            )

        if isinstance(excel_writer, str):
            excel_writer = self.ExcelWriter(excel_writer)
        elif 'openpyxl' not in excel_writer.engine:
            raise ValueError('Engine for excel_writer must be openpyxl.')

        if _os.path.splitext(excel_writer.path)[1] not in excel_writer.supported_extensions:
            raise ValueError(
                'Unsupported file extension {}. Use {}.'.format(
                    _os.path.splitext(excel_writer.path)[1], '/'.join(excel_writer.supported_extensions)
                )
            )

        if replace_sheet:
            if sheet_name in excel_writer.book:
                del excel_writer.book[sheet_name]
            if sheet_name in excel_writer.sheets:
                del excel_writer.sheets[sheet_name]

        self.dataframe.to_excel(
            excel_writer, sheet_name=sheet_name, engine=engine, header=header,
            index=index, startcol=startcol, startrow=startrow, columns=columns, **kwargs
        )

        book = excel_writer.book
        sheet = book[sheet_name]
        sheet.sheet_view.rightToLeft = right_to_left

        # add named styles. Rename any whose name is already taken within book.
        renamed_styles = self._add_named_styles(book)

        if auto_fit is not None and auto_fit is not False:
            if auto_fit is True:
                auto_fit = self.dataframe.columns
            self.auto_fit(auto_fit, index=index, include_header=bool(header))

        # index styles
        if index:
            current_cell = sheet.cell(row=startrow + 1, column=startcol + 1)
            if header and not index_label and self._table_args:
                index_label = 'index'
                current_cell.value = index_label  # Otherwise formatting as table will auto give it a ColumnX name.
            if header:  # TODO: Style this cell properly
                current_cell.style = self._header_styles.iat[0]
            offset = 2 if header else 1
            for row_index, index_style in enumerate(self._index_styles.iteritems()):
                index_value, style = index_style
                current_cell = sheet.cell(row=row_index + startrow + offset, column=startcol + 1)
                current_cell.style = renamed_styles.get(style, style)
            # set index width
            sheet.column_dimensions[self.get_column_letter(startcol)].width = self._index_width
            # adjust startcol for added index column
            startcol += 1

        # header styles
        if header:
            for col_index, col_style in enumerate(self._header_styles.iteritems()):
                col_name, style = col_style
                current_cell = sheet.cell(row=startrow + 1, column=col_index + startcol + 1)
                current_cell.style = renamed_styles.get(style, style)
            # set header height
            sheet.row_dimensions[startrow + 1].height = self.header_height
            # adjust startrow for header row
            startrow += 1

        # data styles
        for col_index, col_series in enumerate(self._styleframe.iteritems()):
            col_name, column = col_series
            for row_index, index_style in enumerate(column.iteritems()):
                index_value, style = index_style
                current_cell = sheet.cell(row=row_index + startrow + 1, column=col_index + startcol + 1)
                current_cell.style = renamed_styles.get(style, style)

        # add any hyperlinks
        if self._hyperlinks is not None:
            for col_index, col_series in enumerate(self._hyperlinks.iteritems()):
                col_name, column = col_series

                try:
                    col_index = self.columns.get_loc(col_name)
                except KeyError:  # col_name not in dataframe
                    if index and col_name in (index_label, self.index.name, 'index'):
                        col_index = -1
                    else:
                        continue

                make_copy = len(column) != column.nunique()
                for row_index, index_link in enumerate(column.iteritems()):
                    index_value, hyperlink = index_link
                    if make_copy and isinstance(hyperlink, _Hyperlink):
                        hyperlink = _copy(hyperlink)

                    current_cell = sheet.cell(row=row_index + startrow + 1, column=col_index + startcol + 1)
                    current_cell.hyperlink = hyperlink

        # set column widths
        for col_index, column_width in enumerate(self._column_widths.iteritems()):
            column, width = column_width
            column_letter = self.get_column_letter(col_index, startcol=startcol)
            sheet.column_dimensions[column_letter].width = width

        # set row heights
        for row_index, row_height in enumerate(self._row_heights.iteritems()):
            row, height = row_height
            sheet.row_dimensions[startrow + row_index + 1].height = height

        # format as table if needed
        if self._table_args:
            self._table_args['ref'] = self._get_range_as_str(startcol=startcol, startrow=headerrow, index=index)

            if not self._table_args['displayName']:  # find next available table name
                tables = {tbl.name for sht in book.worksheets for tbl in sht._tables}
                for i in _count(1):
                    if 'Table{}'.format(i) not in tables:
                        self._table_args['displayName'] = 'Table{}'.format(i)
                        break

            tbl = _table.Table(
                **self._table_args
            )
            sheet.add_table(tbl)

        elif add_filters:
            sheet.auto_filter.ref = self._get_range_as_str(
                row_index=0, startcol=startcol, startrow=headerrow, index=index
            )

        # Hide columns
        if columns_to_hide:
            if isinstance(columns_to_hide, (str, int)):
                column_letter = self.get_column_letter(columns_to_hide, startcol=startcol)
                sheet.column_dimensions[column_letter].hidden = True
            else:
                for column in columns_to_hide:
                    column_letter = self.get_column_letter(column, startcol=startcol)
                    sheet.column_dimensions[column_letter].hidden = True

        # Protect sheet
        if protect_sheet:
            sheet.protection.autoFilter = False
            sheet.protection.enable()

        if save:
            excel_writer.save()

        return excel_writer

    def _add_named_styles(self, book):
        """
        Add named styles for frame to existing workbook.
        Rename any that already exist within workbook.

        :param book: openpyxl workbook
        :return: dict mapping old name: new name for styles that had to be renamed.
        """
        new_styles = dict()
        existing_styles = None
        for name, style in self.named_styles.items():
            try:
                book.add_named_style(style)
            except ValueError:  # Style Exists
                if existing_styles is None:
                    existing_styles = self.named_styles.copy()
                    existing_styles.update({s.name: s for s in book._named_styles})
                if not self._style_eq(existing_styles[name], self.named_styles[name]):
                    s = _copy(style)
                    existing_styles[self._rename(s, existing_styles)] = s
                    book.add_named_style(s)
                    new_styles[name] = s.name
        return new_styles

    def _style_by_type(self, idxr=None, index=False, default_style=None, number_style=None,
                       date_style=None, datetime_style=None, timedelta_style=None):
        """

        :param idxr: (indexes, columns)
        :type idxr: tuple
        :param default_style:
        :param number_style:
        :param date_style:
        :param datetime_style:
        :param timedelta_style:
        :return: None
        """
        idxr = _Slicer._idxr_for_frame(idxr)
        frame = self.dataframe.loc[idxr[0], idxr[1]]
        styles = self.styles[idxr[0], idxr[1]]

        if frame.empty:
            return

        if index:
            frame = frame.reset_index()
            index_name = frame.columns[0]
            styles[index_name] = ''

        if default_style:
            styles.loc[:, :] = self._style_parser(default_style)

        if date_style or datetime_style:
            datetimes = frame.select_dtypes(include=['datetime', 'datetimetz']).columns
            date_cols = [
                col for col in datetimes if self._is_date_col(frame[col])
            ]
            datetime_cols = [
                e for e in datetimes if e not in date_cols
            ]

            if date_style and date_cols:
                styles.loc[:, date_cols] = self._style_parser(date_style)
            if datetime_style and datetime_cols:
                styles.loc[:, datetime_cols] = self._style_parser(datetime_style)

        if number_style:
            styles.loc[:, frame.select_dtypes(include='number').columns] = self._style_parser(number_style)

        if timedelta_style:
            styles.loc[:, frame.select_dtypes(include='timedelta').columns] = self._style_parser(timedelta_style)

        if index:
            self.index_styles.loc[idxr[0]] = self._style_parser(
                _Style._default_index_style(self._named_styles[styles.pop(index_name).iat[0]])
            )

        self._styleframe.loc[idxr[0], idxr[1]] = styles

    @staticmethod
    def _is_date_col(column):
        """
        Check if column is datetime and no entries have a time component.

        :param column: pandas.Series
        :return: boolean
        """
        try:
            return (column.loc[column.notnull()].dt.time != _no_time).sum() == 0
        except AttributeError:
            return False

    def auto_fit(self, columns=None, scalar=None, flat=None, max_width=None, min_width=None,
                 index=True, include_header=True):
        """
        Attempt to auto fit column widths. ~Max length entry in column * scalar + flat.
        If columns not provided fits all columns.

        :param columns: columns to autofit.
        :type columns: list-like.
        :param scalar: to multiply by number of characters to get width. Default utils.Options.default_autofit_scalar.
        :type scalar: float.
        :param flat: flat amount to add to width. Default utils.Options.default_autofit_flat.
        :type flat: float.
        :param max_width: max allowed width. Default utils.Options.default_autofit_max.
        :type max_width: float.
        :param min_width: min allowed width. Default utils.Options.default_autofit_min.
        :type min_width: float.
        :param index: fit index width as well.
        :type index: boolean.
        :param include_header: also consider width of column header when fitting. for index uses index.name.
        :type include_header: boolean.
        :return: self
        """
        if columns is None:
            columns = self._styleframe.columns
        if flat is None:
            flat = _utils.Options.default_autofit_flat
        if scalar is None:
            scalar = _utils.Options.default_autofit_scalar
        if max_width is None:
            max_width = _utils.Options.default_autofit_max
        if min_width is None:
            min_width = _utils.Options.default_autofit_min

        empty_dataframe = self.dataframe.empty
        number_formats = self._styleframe.applymap(lambda style: self._named_styles[style].number_format)

        def fit_column(column, formats):
            dtype = column.dtype.name
            if isinstance(formats, _pd.Series):
                formats = formats.unique()
            format_len = len(max(formats, key=len)) if not empty_dataframe else 0

            if empty_dataframe:
                width = 0
            elif 'date' in dtype and _utils.NumberFormats.general not in formats:
                width = format_len
            elif 'time' in dtype and _utils.NumberFormats.general not in formats:
                width = format_len + len(str(int(column.dt.days.max())))
            elif 'float' in dtype:
                if all('0.0' in e or e == '0' for e in formats):
                    width = format_len + column.round().apply(str).str.len().max() - 2
                else:
                    width = column.apply('{:,.10f}'.format).str.rstrip('0').str.len().max()
            else:
                width = column.apply(str).str.len().max()

            if include_header and column.name:
                width = max(width, len(column.name) + 2)

            return max(min(width * scalar + flat, max_width), min_width)

        for column in columns:
            self._column_widths.at[column] = fit_column(self.dataframe[column], number_formats[column])

        if index:
            index_formats = self._index_styles.apply(lambda style: self._named_styles[style].number_format)
            self._index_width = fit_column(self.index.to_series(), index_formats)

        return self

    def format_as_table(self, table_style=None, table_name=None, row_stripes=True, col_stripes=None, **kwargs):
        """
        When exported format excel range as a table. Tables names must be unique within a workbook.
        To format as table must use headers and filters will be enabled.

        :param table_style: Table style to use.
        :type table_style: string (Ex. 'TableStyleLight1') or openpyxl.worksheet.table.TableStyleInfo
        :param table_name: Name for table. Must be unique within workbook.
            Default will find the next available name counting up. Table1, Table2, Table3 ect.
        :type table_name: string
        :param row_stripes: Show row stripes.
        :type row_stripes: boolean
        :param col_stripes: Show col stripes.
        :type col_stripes: boolean
        :param kwargs: Passed to openpyxl.worksheet.table.Table
        :return: self
        """

        if isinstance(table_style, str):
            table_style = _table.TableStyleInfo(
                name=table_style,
                showRowStripes=row_stripes,
                showColumnStripes=col_stripes,
            )
        kwargs['tableStyleInfo'] = table_style
        kwargs['displayName'] = table_name
        self._table_args = kwargs
        return self

    def clear_table_formatting(self):
        """
        Remove table formatting.

        :return: self
        """
        self._table_args = None
        return self

    def row_stripes(self, fill_color='D9D9D9'):
        """
        Solid fill every other row with fill_color.

        :param fill_color: Color to fill. Hex, rgb tuple or or openpyxl.styles.Color.
        :type fill_color: str, tuple or openpyxl.styles.Color
        :return: self
        """
        _filler['fill_color'] = fill_color
        self.styles[::2, :] = _filler
        return self

    def col_stripes(self, fill_color='D9D9D9'):
        """
        Solid fill every other column with fill_color.

        :param fill_color: Color to fill. Hex, rgb tuple or or openpyxl.styles.Color.
        :type fill_color: str, tuple or openpyxl.styles.Color
        :return: self
        """
        _filler['fill_color'] = fill_color
        self.styles[:, ::2] = _filler
        return self

    def get_column_letter(self, column, startcol=0):
        """
        Get excel column letter for a given frame column.

        :param column: column name or number. column numbers 0 based.
        :type column: str/int
        :param startcol: column offset
        :type startcol: int
        :return: column letter
        """

        if not isinstance(column, (int, str)):
            raise TypeError("Column must be an index or column name.")

        idx = None
        # worksheet columns index start from 1
        if column in self.dataframe.columns:  # column name
            idx = self.dataframe.columns.get_loc(column) + startcol + 1
        elif isinstance(column, int) and column >= 0:  # column index
            idx = column + startcol + 1

        if idx is None:
            raise IndexError("Column {} is out of columns range.".format(str(column)))

        return _get_column_letter(idx)

    def add_style(self, style):
        """
        Add style to available named styles. Can be assigned just by name afterwards.
        Styles will also be automatically added when first assigned.

        :param style: Style to add
        :type style: xlframe.Style or openpyxl.styles.NamedStyle
        :return: Style name
        :rtype: str
        """
        if style.name in self._named_styles:
            if self._style_eq(style, self._named_styles[style.name]):
                return style.name
            raise KeyError('Style by name "{}" already exists'.format(style.name))

        return self._style_parser(style)

    def _style_parser(self, style):
        """
        Sort out different acceptable style arguments.

        :param style: xlframe.Style, openpyxl.styles.NamedStyle or str
        :return: Style name
        :rtype: str
        """
        if isinstance(style, str):
            if style not in self._named_styles:
                raise KeyError('Style by name {} not found.'.format(style))
            return style

        if isinstance(style, _Style):
            style = style.named_style

        if style.name in self._named_styles:
            if not self._style_eq(style, self._named_styles[style.name]):
                raise KeyError('Style by name {} already exists'.format(style.name))
            return style.name

        self._add_style(style)
        return style.name

    def _add_style(self, style):
        """
        Add style to appropriate containers.

        :param style: openpyxl.style.NamedStyle
        :return: None
        """
        self.named_styles[style.name] = style
        self._named_styles[style.name] = style

    def _style_editor(self, idxr, source, changes):
        """
        Make changes to existing styles of source at idxr.

        :param idxr: (index, columns) to apply changes to.
        :param source: source dataframe.loc to index with idxr
        :type source: pandas.DataFrame.loc
        :param changes: Changes to apply to styles of source at idxr. kwargs for xlframe.Style.
        :type changes: Dict
        :return: List of newly created names.
        :rtype: List of strings
        """
        cache = dict()
        changes = tuple(changes.items())
        section = source[idxr]

        if isinstance(section, _pd.DataFrame):
            source[idxr] = section.applymap(lambda style: self._style_edit(style, changes=changes, cache=cache))
        elif isinstance(section, _pd.Series):
            source[idxr] = section.apply(self._style_edit, changes=changes, cache=cache)
        else:
            source[idxr] = self._style_edit(section, changes=changes, cache=cache)
        return [style_name for _, style_name in cache.items()]

    def _style_edit(self, style_name, changes, cache=None):
        """
        Make changes to style style_name and register it with frame.

        :param style_name: Name of style
        :type style_name: str
        :param changes: Changes to make. kwargs for xlframe.Style
        :type changes: Dict
        :param cache: Optional cache of already edited styles to avoid recreating the same edit.
        :type cache: Dict
        :return: Style name
        :rtype: str
        """
        if not changes:
            return style_name
        if cache:
            try:
                return cache[(style_name, changes)]
            except KeyError:
                pass

        s = _Style(self._named_styles[style_name])

        named = False
        for attr, value in changes:
            setattr(s, attr, value)
            if attr == 'name':
                named = True

        if named:
            if s.name in self._named_styles and self._style_eq(s, self._named_styles[s.name]):
                if cache is not None:
                    cache[(style_name, changes)] = s.name
                return s.name
            elif s.name not in self._named_styles:
                if cache is not None:
                    cache[(style_name, changes)] = s.name
                return self.add_style(s)
            raise ValueError(
                'More than one unique style created. '
                'Cannot use name "{}" again.'.format(s.name)
            )

        if cache is not None:
            cache[(style_name, changes)] = self._rename(s)

        return self.add_style(s)

    def _rename(self, style, named_styles=None):
        """
        Give styles a new (numbered) name. Finds next available name.

        :param style: Style to rename.
        :type style: openpyxl.style.NamedStyle or xlframe.Style
        :param named_styles: Optional dict of existing styles. Uses frames available styles if unspecified.
        :return: New name
        :rtype: str
        """
        if named_styles is None:
            named_styles = self._named_styles
        style_name, style_num = self._find_name_num(style.name)

        for i in _count(style_num + 1):
            new_name = '{}[{}]'.format(style_name, i)
            if new_name not in named_styles:
                style.name = new_name
                return new_name

    @staticmethod
    def _find_name_num(style_name):
        if style_name[-1] != ']':
            return style_name, 0
        try:
            return style_name[:style_name.rindex('[')], int(style_name[style_name.rindex('[') + 1:-1])
        except ValueError:  # substring not found or invalid input to int() found.
            return style_name, 0

    def _get_range_as_str(self, row_index=None, columns=None, startcol=0, startrow=0, index=False):
        """
        Get string representation of cell range. 'A1:B4'

        :param row_index: row number/s to include. default includes all dataframe rows
        :type row_index: int/s
        :param columns: dataframe columns to include. default includes all dataframe columns.
        :type columns: list-like of strings
        :param startcol: column offset
        :type startcol: int
        :param startrow: row offset
        :type startrow: int
        :param index:
        :type index: bool
        :return: String representation of cell range. ex 'A1:E10'
        :rtype: str
        """
        if columns is None:  # returns cells range for all columns
            start_letter = self.get_column_letter(self.dataframe.columns[0], startcol=startcol - index)
            end_letter = self.get_column_letter(self.dataframe.columns[-1], startcol=startcol)
        else:
            if isinstance(columns, (int, str)):
                start_letter = self.get_column_letter(columns, startcol=startcol)
                end_letter = start_letter
            else:
                start_letter = self.get_column_letter(columns[0], startcol=startcol)
                end_letter = self.get_column_letter(columns[-1], startcol=startcol)

        if row_index is None:  # returns cells range for all rows
            start_index = startrow + 1
            end_index = start_index + len(self.dataframe)
        else:
            try:
                start_index = startrow + row_index[0] + 1
                end_index = start_index + row_index[-1]
            except TypeError:  # single row
                start_index = startrow + row_index + 1
                end_index = start_index

        return '{start_letter}{start_index}:{end_letter}{end_index}'.format(
            start_letter=start_letter, start_index=start_index, end_letter=end_letter, end_index=end_index
        )

    def _style_eq(self, style1, style2):
        """
        Check if 2 styles are equal.

        :param style1:
        :param style2:
        :return:
        """
        if isinstance(style1, str):
            style1 = self._named_styles[style1]
        if isinstance(style2, str):
            style2 = self._named_styles[style2]
        if isinstance(style1, _Style):
            return style1 == style2
        if isinstance(style2, _Style):
            return style2 == style1
        return hash(style1) == hash(style2)

    def __getitem__(self, item):
        return self.dataframe.__getitem__(item)

    def __setitem__(self, key, style):
        if isinstance(style, dict):
            return self._style_editor(key, self._styleframe, style)
        style = self._style_parser(style)
        return self._styleframe.__setitem__(key, style)

    def __delitem__(self, key):
        raise NotImplementedError

    def __str__(self):
        s = '{}\n\n{}\n\n{}\n\n{}\n\n{}\n\n{}'
        return s.format(
            str(self.dataframe),
            str(self._styleframe),
            str(self._header_styles),
            str(self._index_styles),
            str(self._column_widths),
            str(self._row_heights),
        )

    def __len__(self):
        return len(self._styleframe)

    def __iter__(self):
        return iter(self._styleframe)

    @property
    def styles(self):
        return self._styleframe.style_loc

    @styles.setter
    def styles(self, style):
        self.styles[:, :] = style

    @property
    def istyles(self):
        return self._styleframe.style_iloc

    @istyles.setter
    def istyles(self, style):
        self.istyles[:, :] = style

    @property
    def loc(self):
        return self._slicer

    @property
    def iloc(self):
        return self._islicer

    @property
    def index(self):
        return self._styleframe.index

    @property
    def columns(self):
        return self._styleframe.columns

    @property
    def header_styles(self):
        return self._header_styles.style_idxr

    @header_styles.setter
    def header_styles(self, style):
        self.header_styles[:] = style

    @property
    def index_styles(self):
        return self._index_styles.style_idxr

    @index_styles.setter
    def index_styles(self, style):
        self.index_styles[:] = style

    @property
    def row_heights(self):
        return self._row_heights

    @row_heights.setter
    def row_heights(self, value):
        self.row_heights[:] = value

    @property
    def column_widths(self):
        return self._column_widths

    @column_widths.setter
    def column_widths(self, value):
        self.column_widths[:] = value

    @property
    def header_height(self):
        return self._header_height

    @header_height.setter
    def header_height(self, value):
        self._header_height = float(value)

    @property
    def index_width(self):
        return self._index_width

    @index_width.setter
    def index_width(self, value):
        self._index_width = float(value)

    @property
    def hyperlinks(self):
        if self._hyperlinks is None:
            self._hyperlinks = _pd.DataFrame(index=self.index)
        return self._hyperlinks

    @staticmethod
    def ExcelWriter(path, load_existing=False, **kwargs):
        """
        See pandas.ExcelWriter. Engine will be set to 'openpyxl'.

        :param path: Full path for workbook.
        :type path: String.
        :param load_existing: Load existing workbook into excel_writer if one exists at path.
        :type load_existing: Boolean.
        :return: pandas.ExcelWriter
        """
        kwargs['engine'] = 'openpyxl'
        kwargs['date_format'] = kwargs.get('date_format', _utils.Options.default_date_format)
        kwargs['datetime_format'] = kwargs.get('datetime_format', _utils.Options.default_datetime_format)
        excel_writer = _pd.ExcelWriter(path, **kwargs)

        if load_existing and _os.path.isfile(path):
            # load book
            vba = _os.path.splitext(path)[1] == '.xlsm'
            book = _load_workbook(path, keep_vba=vba)
            excel_writer.book = book

            # add sheets from book to writer
            for sheet in book:
                excel_writer.sheets[sheet.title] = sheet

        return excel_writer

    def __eq__(self, other):
        attrs = (
            'header_height', 'index_width',
        )
        attrs_pd = (
            'dataframe', '_styleframe',
            '_index_styles', '_header_styles',
            '_row_heights', '_column_widths'
        )
        shared_styles = set(self.named_styles) & set(other.named_styles)
        tbl = other._table_args is None if self._table_args is None else self._table_args == other._table_args
        hypr = other._hyperlinks is None if self._hyperlinks is None else self._hyperlinks.equals(other._hyperlinks)

        return tbl and hypr \
               and all(getattr(self, attr) == getattr(other, attr) for attr in attrs) \
               and all(getattr(self, attr).equals(getattr(other, attr)) for attr in attrs_pd) \
               and all(self._style_eq(self.named_styles[style], other.named_styles[style]) for style in shared_styles)


class _StyleIndexer:
    def __init__(self, styler, indexer):
        self.styler = styler
        self.indexer = indexer

    def __setitem__(self, key, style):
        if isinstance(style, dict):
            self.styler._style_editor(key, self.indexer, style)
        else:
            style = self.styler._style_parser(style)
            self.indexer.__setitem__(key, style)

    def __getitem__(self, item):
        return self.indexer.__getitem__(item)

    def __getattr__(self, item):
        return getattr(self.indexer, item)

    def __str__(self):
        return str(self.indexer.obj)

    def __repr__(self):
        return repr(self.indexer.obj)


class _SeriesIndexer:
    def __init__(self, styler, series):
        self.styler = styler
        self.series = series
        self._loc = _StyleIndexer(styler, series.loc)
        self._iloc = _StyleIndexer(styler, series.iloc)

    @property
    def loc(self):
        return self._loc

    @property
    def iloc(self):
        return self._iloc

    @property
    def ix(self):
        raise NotImplementedError

    def __setitem__(self, key, style):
        if isinstance(style, dict):
            self.styler._style_editor(key, self.series, style)
        else:
            style = self.styler._style_parser(style)
            self.series.__setitem__(key, style)

    def __getitem__(self, item):
        return self.series.__getitem__(item)

    def __getattr__(self, item):
        return getattr(self.series, item)

    def __str__(self):
        return str(self.series)

    def __repr__(self):
        return repr(self.series)


class _Slicer:
    def __init__(self, styler, idx_by='loc'):
        self.styler = styler
        self.idx_by = idx_by

    def __setitem__(self, key, value):
        raise NotImplementedError

    def __getitem__(self, item):
        return self._slice(self.styler, item, self.idx_by)

    @staticmethod
    def _idxr_for_frame(idxr):
        """
        Change idxr to make sure indexing a DataFrame will return a DataFrame
        and indexing a Series will return a Series.
        :param idxr: (indexes, columns)
        :type idxr: tuple
        :return: tuple
        """
        if isinstance(idxr[0], (str, int, float)):
            idxr = ([idxr[0]], idxr[1])
        if isinstance(idxr[1], (str, int, float)):
            idxr = (idxr[0], [idxr[1]])
        return idxr

    @staticmethod
    def _slice(source, idxr, idx_by='loc'):
        """

        :param source: Source XlFrame to slice.
        :type source: xlframe.XlFrame
        :param idxr: (indexes, columns) to get.
        :type idxr: tuple
        :param idx_by: indexing choice. loc or iloc.
        :type idx_by: string
        :return: New XlFrame
        """
        idxr = _Slicer._idxr_for_frame(idxr)

        frame = XlFrame(getattr(source.dataframe, idx_by).__getitem__(idxr), use_default_formats=False)
        frame.named_styles = source.named_styles.copy()
        frame._named_styles = source._named_styles.copy()

        frame._styleframe.loc[:, :] = getattr(source._styleframe, idx_by).__getitem__(idxr).values

        frame._index_styles.loc[:] = getattr(source._index_styles, idx_by)[idxr[0]].values
        frame._header_styles.loc[:] = getattr(source._header_styles, idx_by)[idxr[1]].values

        frame._row_heights.loc[:] = getattr(source._row_heights, idx_by)[idxr[0]].values
        frame._column_widths.loc[:] = getattr(source._column_widths, idx_by)[idxr[1]].values

        frame._index_width = source._index_width
        frame._header_height = source._header_height

        frame._defaults_used = source._defaults_used

        if source._table_args is not None:
            frame._table_args = source._table_args.copy()

        if source._hyperlinks is not None:
            links = getattr(source._hyperlinks, idx_by)[idxr[0], :]
            for col in [c for c in source._hyperlinks if c in frame.columns or c in (frame.index.name, 'index')]:
                frame.hyperlinks[col] = links[col]

        return frame


if __name__ == '__main__':
    pass
