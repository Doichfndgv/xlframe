from copy import copy as _copy

from openpyxl.styles import NamedStyle, PatternFill, Protection, Border, Font, Side, Color, Alignment
from openpyxl.styles.builtins import styles as _styles
from openpyxl.styles.colors import aRGB_REGEX as _aRGB_REGEX, COLOR_INDEX as _COLOR_INDEX
from openpyxl.xml.functions import fromstring as _fromstring, QName as _QName

from . import utils as _utils

__all__ = ['Style']
_theme_builtins = [
    'FFFFFF', '000000', 'EEECE1', '1F497D', '4F81BD', 'C0504D', '9BBB59', '8064A2', '4BACC6', 'F79646'
]


def _color_parser(color):
    if color is None:
        return
    elif isinstance(color, Color):
        return _copy(color)
    elif isinstance(color, tuple):
        color = _rgb_to_hex(color)

    if isinstance(color, str):
        try:
            color = getattr(_utils.Colors, color.lower())
        except AttributeError:
            pass

        if color.startswith('#'):
            color = color[1:]
        if _aRGB_REGEX.match(color):
            return Color(
                rgb=color
            )

    raise ValueError(
        'Invalid color {}.\n\nMust be (r, g, b) tuple, hex string, '
        'utils.Colors name or openpyxl Color.'.format(str(color))
    )


def _get_theme_colors(book):
    # https://groups.google.com/forum/#!msg/openpyxl-users/v2FDsbDDTqU/rQWLAXZFkeUJ
    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = _fromstring(book.loaded_theme)
    themeEl = root.find(_QName(xlmns, 'themeElements').text)
    colorSchemes = themeEl.findall(_QName(xlmns, 'clrScheme').text)
    firstColorScheme = colorSchemes[0]
    colors = []
    for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
        accent = firstColorScheme.find(_QName(xlmns, c).text)

        if 'window' in accent.getchildren()[0].attrib['val']:
            colors.append(accent.getchildren()[0].attrib['lastClr'])
        else:
            colors.append(accent.getchildren()[0].attrib['val'])
    return colors


def _rgb_to_hex(*args):
    if len(args) > 1:
        r, g, b = args
    else:
        r, g, b = args[0]
    return "{:02x}{:02x}{:02x}".format(int(r), int(g), int(b))


def _hex_to_rgb(color):
    if color.startswith('#'):
        color = color[1:]
    if _aRGB_REGEX.match(color):
        color = color[-6:]
        return tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
    raise ValueError('Invalid hex code {}.'.format(color))


def _hex_from_theme(color, theme_colors=None):
    if not theme_colors:
        theme_colors = _theme_builtins
    r, g, b = _hex_to_rgb(theme_colors[color.theme])
    r = r + (255 - r) * color.tint
    g = g + (255 - g) * color.tint
    b = b + (255 - b) * color.tint
    return _rgb_to_hex(r, g, b)


class Style:

    __slots__ = (
        'name', 'number_format', 'font_style', 'font_size', '_font_color', 'bold', 'underline', 'italic',
        'strikethrough', 'fill_pattern', '_fill_color', 'horizontal_alignment', 'vertical_alignment', 'indent',
        'wrap_text', 'shrink_to_fit', 'locked', 'hidden', 'left', 'right', 'top', 'bottom'
    )

    def __init__(self, name, number_format=_utils.NumberFormats.general, font_style=_utils.Options.default_font_style,
                 font_size=_utils.Options.default_font_size, font_color=None, bold=None, underline=None, italic=None,
                 strikethrough=None, fill_pattern='solid', fill_color=None, horizontal_alignment=None,
                 vertical_alignment=None, indent=0, wrap_text=None, shrink_to_fit=None,
                 border_style=None, border_color=None):
        """
        Optional constructor for styles.
        Style.to_named_style() or Style.named_style to get equivalent openpyxl.styles.NamedStyle.
        Style.from_named_style(named_style) to convert openpyxl.styles.NamedStyle to equivalent Style.

        :param name: Style name.
        :type name: str
        :param number_format: Excel number format. See utils.NumberFormats.
        :type number_format: str
        :param font_style: Excel font style. See utils.FontStyles.
        :type font_style: str
        :param font_size: Font size.
        :type font_size: int/float
        :param font_color: Font color. See utils.Colors.
        :type font_color: Hex str, (r, g, b) tuple or openpyxl.styles.Color
        :param bold: Make font bold.
        :type bold: boolean
        :param underline: Underline font. See utils.Underline.
        :type underline: boolean
        :param italic: Italic font.
        :type italic: boolean
        :param strikethrough: strikethrough font.
        :type strikethrough: boolean
        :param fill_pattern: Excel fill pattern. See utils.FillPattern.
        :type fill_pattern: str
        :param fill_color: Fill color. See utils.Colors.
        :type fill_color: Hex str, (r, g, b) tuple or openpyxl.styles.Color
        :param horizontal_alignment: Excel horizontal alignment. See utils.Alignments.Horizontal.
        :type horizontal_alignment: str
        :param vertical_alignment: Excel vertical alignment. See utils.Alignments.Vertical.
        :type vertical_alignment: str
        :param indent: Cell indent.
        :type indent: int
        :param wrap_text: Enable wrap text.
        :type wrap_text: boolean
        :param shrink_to_fit: Enable shrink to fit.
        :type shrink_to_fit: boolean
        :param border_style: Border style. See utils.BorderStyles.
        :type border_style: str
        :param border_color: Border color. See utils.Colors.
        :type border_color: Hex str, (r, g, b) tuple or openpyxl.styles.Color
        """

        # (border_style, border_color)
        self.left = (None, None)
        self.right = (None, None)
        self.top = (None, None)
        self.bottom = (None, None)

        if isinstance(name, NamedStyle):
            self.named_style = name
        else:
            self.name = name
            self.number_format = number_format
            self.font_style = font_style
            self.font_size = float(font_size)
            self.font_color = font_color
            self.bold = bold
            self.underline = underline
            self.italic = italic
            self.strikethrough = strikethrough
            self.fill_pattern = fill_pattern
            self.fill_color = fill_color
            self.horizontal_alignment = horizontal_alignment
            self.vertical_alignment = vertical_alignment
            self.indent = float(indent)
            self.wrap_text = wrap_text
            self.shrink_to_fit = shrink_to_fit

            self.border_style = border_style
            self.border_color = border_color

            self.locked = True
            self.hidden = False

    @property
    def font_color(self):
        return self._font_color

    @font_color.setter
    def font_color(self, color):
        self._font_color = _color_parser(color)

    @property
    def fill_color(self):
        return self._fill_color

    @fill_color.setter
    def fill_color(self, color):
        self._fill_color = _color_parser(color)

    @property
    def named_style(self):
        return NamedStyle(
            name=self.name,
            font=self.font,
            fill=self.fill,
            alignment=self.alignment,
            border=self.border,
            number_format=self.number_format,
            protection=self.protection
        )

    @named_style.setter
    def named_style(self, named_style):
        self.name = named_style.name
        self.number_format = named_style.number_format
        self.font = named_style.font
        self.fill = named_style.fill
        self.alignment = named_style.alignment
        self.border = named_style.border
        self.protection = named_style.protection

    @property
    def sides(self):
        return tuple(
            Side(
                border_style=e[0],
                color=e[1]
            ) if e[0] else Side() for e in self._sides
        )

    @property
    def _sides(self):
        return self.left, self.right, self.top, self.bottom

    @property
    def border_style(self):
        return tuple(e[0] for e in self._sides)

    @border_style.setter
    def border_style(self, style):
        self.left = (style, self.left[1])
        self.right = (style, self.right[1])
        self.top = (style, self.top[1])
        self.bottom = (style, self.bottom[1])

    @property
    def border_color(self):
        return tuple(e[1] for e in self._sides)

    @border_color.setter
    def border_color(self, color):
        self.left = (self.left[0], _color_parser(color))
        self.right = (self.right[0], _color_parser(color))
        self.top = (self.top[0], _color_parser(color))
        self.bottom = (self.bottom[0], _color_parser(color))

    @property
    def font(self):
        return Font(
            name=self.font_style,
            size=self.font_size,
            color=self.font_color,
            bold=self.bold,
            underline=self.underline,
            strikethrough=self.strikethrough,
            italic=self.italic
        )

    @font.setter
    def font(self, font):
        self.font_style = font.name
        self.font_size = font.sz
        self.font_color = _color_parser(font.color)
        self.bold = font.b
        self.italic = font.i
        self.underline = font.u
        self.strikethrough= font.strike

    @property
    def fill(self):
        if self.fill_color:
            return PatternFill(
                fill_type=self.fill_pattern,
                fgColor=self.fill_color,
                bgColor=Color()
            )
        return PatternFill()

    @fill.setter
    def fill(self, pattern_fill):
        self.fill_color = _color_parser(pattern_fill.fgColor)
        self.fill_pattern = pattern_fill.patternType

    @property
    def alignment(self):
        return Alignment(
            horizontal=self.horizontal_alignment,
            vertical=self.vertical_alignment,
            wrap_text=self.wrap_text,
            shrink_to_fit=self.shrink_to_fit,
            indent=self.indent
        )

    @alignment.setter
    def alignment(self, alignment):
        self.horizontal_alignment = alignment.horizontal
        self.vertical_alignment = alignment.vertical
        self.wrap_text = alignment.wrapText
        self.shrink_to_fit = alignment.shrinkToFit
        self.indent = alignment.indent

    @property
    def border(self):
        return Border(
            *self.sides
        )

    @border.setter
    def border(self, border):
        self.left = border.left.style, _color_parser(border.left.color)
        self.right = border.right.style, _color_parser(border.right.color)
        self.top = border.top.style, _color_parser(border.top.color)
        self.bottom = border.bottom.style, _color_parser(border.bottom.color)

    @property
    def protection(self):
        return Protection(
            locked=self.locked,
            hidden=self.hidden
        )

    @protection.setter
    def protection(self, protection):
        self.locked = protection.locked
        self.hidden = protection.hidden

    def to_named_style(self):
        """
        Convert to equivalent openpyxl.styles.NamedStyle.

        :return: openpyxl.styles.NamedStyle
        """
        return self.named_style

    @classmethod
    def from_named_style(cls, style):
        """
        Convert openpyxl.styles.NamedStyle to equivalent Style.

        :param style: openpyxl.styles.NamedStyle
        :return: Style
        """
        return cls(style)

    @staticmethod
    def opxl_color_to_rgb(color, book=None, theme_colors=None):
        return _hex_to_rgb(Style.opxl_color_to_hex(color, book, theme_colors))

    @staticmethod
    def opxl_color_to_hex(color, book=None, theme_colors=None):
        if color.type == 'rgb' and _aRGB_REGEX.match(color.rgb):
            return color.rgb[-6:]
        elif color.type == 'theme':
            if book:
                theme_colors = _get_theme_colors(book)
            if not theme_colors:
                theme_colors = _theme_builtins
            return _hex_from_theme(color, theme_colors)[-6:]
        elif color.type == 'indexed':
            # TODO: Color index 64/65 are not hex values.
            return _COLOR_INDEX[color.indexed][-6:]
        else:
            raise ValueError('Error parsing color:\n{}'.format(str(color)))

    @classmethod
    def default_style(cls):
        default = cls(_styles['Normal'])
        default.name = 'Default'
        default.font_style = _utils.Options.default_font_style
        default.font_size = _utils.Options.default_font_size
        return default.named_style

    @classmethod
    def _default_header_style(cls, default=None):
        if default is None:
            default = cls.default_style()
        default = cls(default)
        default.name += ' Header'
        default.bold = True
        default.border_style = _utils.BorderStyles.thin
        default.number_format = _utils.NumberFormats.text
        default.horizontal_alignment = _utils.Alignments.Horizontal.center
        return default.named_style

    @classmethod
    def _default_index_style(cls, default=None):
        if default is None:
            default = cls.default_style()
        default = cls(default)
        default.name += ' Index'
        default.bold = True
        default.border_style = _utils.BorderStyles.thin
        return default.named_style

    @classmethod
    def _default_number_style(cls, default=None):
        if default is None:
            default = cls.default_style()
        default = cls(default)
        default.name += ' Number'
        default.horizontal_alignment = _utils.Alignments.Horizontal.right
        default.number_format = _utils.NumberFormats.general
        return default.named_style

    @classmethod
    def _default_date_style(cls, default=None):
        if default is None:
            default = cls.default_style()
        default = cls(default)
        default.name += ' Date'
        default.horizontal_alignment = _utils.Alignments.Horizontal.right
        default.number_format = _utils.Options.default_date_format
        return default.named_style

    @classmethod
    def _default_datetime_style(cls, default=None):
        if default is None:
            default = cls.default_style()
        default = cls(default)
        default.name += ' Datetime'
        default.horizontal_alignment = _utils.Alignments.Horizontal.right
        default.number_format = _utils.Options.default_datetime_format
        return default.named_style

    @classmethod
    def _default_timedelta_style(cls, default=None):
        if default is None:
            default = cls.default_style()
        default = cls(default)
        default.name += ' Timedelta'
        default.number_format = _utils.Options.default_timedelta_format

        if any(e in default.number_format for e in ('year', 'day', 'hour', 'min', 'sec')):
            default.horizontal_alignment = _utils.Alignments.Horizontal.left
        else:
            default.horizontal_alignment = _utils.Alignments.Horizontal.right

        return default.named_style

    def as_dict(self):
        raise NotImplementedError

    def as_tuple(self):
        return tuple(
            (attr, getattr(self, attr)) for attr in self.__slots__
        )

    def __hash__(self):
        return hash(self.as_tuple())

    def __eq__(self, other):
        if isinstance(other, NamedStyle):
            return self == Style(other)
        return self.as_tuple() == other.as_tuple()


if __name__ == '__main__':
    pass
